import csv
import io
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook


def parse_csv(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    columns = [c.strip() for c in (reader.fieldnames or [])]
    rows = [dict(row) for row in reader]
    return columns, rows


def parse_excel(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header = next(rows_iter, None) or ()
    columns = [str(col).strip() if col is not None else "" for col in header]
    rows: List[Dict[str, Any]] = []
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue
        row = {columns[i]: values[i] for i in range(len(columns)) if i < len(values)}
        rows.append(row)
    return columns, rows


def parse_datasheet_file(filename: str, content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    lowered = (filename or "").lower()
    if lowered.endswith(".csv"):
        return parse_csv(content)
    if lowered.endswith(".xlsx") or lowered.endswith(".xlsm"):
        return parse_excel(content)
    raise ValueError("Unsupported file type. Upload a .csv or .xlsx file.")


def validate_columns(columns: List[str], expected_fields: List[str]) -> None:
    upload_set = {c.strip() for c in columns if c and c.strip()}
    expected_set = {f.strip() for f in expected_fields if f and f.strip()}
    missing = expected_set - upload_set
    if missing:
        raise ValueError(
            "Uploaded file does not match the template. Missing required columns: "
            f"{', '.join(sorted(missing))}. Template expects: {', '.join(sorted(expected_set))}."
        )
